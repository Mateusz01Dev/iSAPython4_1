# założenie - często pracujemy z plikami excel, zmieniamy w nich jakieś
# dane, i zapisujemy kopię
# otwarcie - jest na początku pracy
# zapis i zamknięcie pliku na końcu

from contextlib import contextmanager
import openpyxl

# bez context manadżera
plik = openpyxl.load_workbook("example.xlsx")

# coś robimy
# ...
# czyszczenie - zamykanie - czynności, które robimy na koniec pracy
plik.save("example-copy.xlsx")
plik.close()


# z context managerem

@contextmanager
def otworz_excel(plik):
    # ten kod jest wykonywany na wejściu do bloku with
    xlsx = openpyxl.load_workbook(plik)

    # yield - zwracamy jakiś obiekt
    # jeśli nie chcemy zwracać żadnego obiektu to możemy
    # dać samo yield
    yield xlsx

    # ten kod będzie wykonay na wyjściu z bloku with
    nowa_nazwa = "example-copy.xlsx"
    xlsx.save(nowa_nazwa)


# wchodzimy do bloku więc wykonane zostaną linijki 24, 29
with otworz_excel("example.xlsx") as plik_excel:
    # coś robimy
    print(plik_excel.sheetnames)

# wyszliśmy z bloku with, teraz wykonane są linijki 32, 33
